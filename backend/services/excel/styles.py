from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from accounting.rules import PL_SUBTOTAL_LABELS, BS_SUBTOTAL_LABELS, BS_PARTIDA_LABELS
from accounting.aggregation import TOTAL_COL
from config.calendar import MONTH_NAMES_SET


# ── Global Font ──────────────────────────────────────────────────────────────
# Calibri 11pt, black — applied to every cell
DEFAULT_FONT = Font(name="Calibri", size=11, color="000000")

# ── Number formats ───────────────────────────────────────────────────────────
NUM_FMT = '#,##0;(#,##0);-'

# ── Title style (B2) ────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="000000")

# ── Header row style (month labels row) ─────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_BORDER = Border(bottom=Side(style="thin", color="000000"))
HEADER_CENTER = Alignment(horizontal="center")

# ── Bold (for TOTAL columns/rows) ───────────────────────────────────────────
BOLD = Font(name="Calibri", size=11, bold=True, color="000000")

# ── No fill / no border (for clearing) ──────────────────────────────────────
NO_FILL   = PatternFill(fill_type=None)
NO_BORDER = Border()

# ── Fills (kept for specific use: excluded rows, PL subtotals) ──────────────
HIGHLIGHT_FILL   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SUBTOTAL_FILL    = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
UNDEFINED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
UNDEFINED_FONT   = Font(name="Calibri", size=11, color="9C0006")                             # Dark red text
DOTTED_BORDER  = Border(
    top=Side(style="dotted"),
    bottom=Side(style="dotted"),
    left=Side(style="dotted"),
    right=Side(style="dotted"),
)

# ── Column widths ────────────────────────────────────────────────────────────
COL_B_WIDTH_DEFAULT = 12
COL_B_WIDTH_PL      = 40
COL_B_WIDTH_BS      = 55

# ── Sheet layout constants ───────────────────────────────────────────────────
TITLE_ROW      = 2
DATA_START_ROW = 3   # 0-indexed for pandas (= Excel row 4 headers)
DATA_START_COL = 1   # 0-indexed for pandas (= Excel column B)
HEADER_ROW     = DATA_START_ROW + 1  # Excel row 4 (1-indexed)
EXCEL_DATA_ROW = 5   # 1-indexed Excel row where first data row appears

# ── Column index constants (1-indexed Excel columns) ────────────────────────
PL_TOTAL_COL           = 15   # Column O — TOTAL for PL sheet
STANDARD_TOTAL_COL     = 16   # Column P — TOTAL for standard detail sheets
WIDE_TOTAL_COL         = 18   # Column R — TOTAL for wide (CC x CC) sheets
STANDARD_FIRST_VAL_COL = 4    # Column D — first value column (standard)
WIDE_FIRST_VAL_COL     = 6    # Column F — first value column (wide)

# Sales details layout (consistent with other sheets: title row 2, headers row 4)
SALES_PANDAS_STARTROW = 3
SALES_HEADER_ROW      = 4

# Gap rows between sections
EXPANDED_GAP_ROWS = 5
PE_GAP_ROWS       = 2

# ── Column index constants (1-indexed, for cell/column references) ───────────
COL_A = 1
COL_B = 2       # first label column
COL_C = 3       # description / first PL value column
COL_D = 4
COL_E = 5

# ── Column width constants ───────────────────────────────────────────────────
COL_A_WIDTH = 5
VAL_COL_WIDTH = 11
DESC_COL_WIDTH = 40
WIDE_LABEL_COL_WIDTH = 12
BS_VAL_COL_WIDTH = 14


# ── Low-level styling helpers ────────────────────────────────────────────────

def set_standard_col_widths(ws):
    """Set column C=40, D-through-P=11 (standard detail/validator layout)."""
    ws.column_dimensions["C"].width = DESC_COL_WIDTH
    for col_idx in range(STANDARD_FIRST_VAL_COL, STANDARD_TOTAL_COL + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = VAL_COL_WIDTH


def _set_wide_col_widths(ws):
    """Set B&D=12, C&E=40, F-through-R=11 (wide CC-x-CC / VALIDADOR3 layout)."""
    ws.column_dimensions["B"].width = WIDE_LABEL_COL_WIDTH
    ws.column_dimensions["C"].width = DESC_COL_WIDTH
    ws.column_dimensions["D"].width = WIDE_LABEL_COL_WIDTH
    ws.column_dimensions["E"].width = DESC_COL_WIDTH
    for col_idx in range(WIDE_FIRST_VAL_COL, WIDE_TOTAL_COL + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = VAL_COL_WIDTH


def _clear_header_cell(ws, row, col):
    """Remove value, border, fill, and bold from a header cell."""
    cell = ws.cell(row=row, column=col)
    cell.value = None
    cell.border = NO_BORDER
    cell.fill = NO_FILL
    cell.font = DEFAULT_FONT


def clear_header_cells(ws, row):
    """Clear the label header cells (columns B and C) for a given row."""
    _clear_header_cell(ws, row, COL_B)
    _clear_header_cell(ws, row, COL_C)


def style_header_row(ws, row, first_value_col, last_col):
    """Apply header styling: bold, gray fill, center on value cols, thin bottom border.

    first_value_col: 1-indexed column where centering starts (value columns).
    Skips cells with no value (already cleared label cells).
    """
    for col in range(COL_B, last_col + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            continue
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        if col >= first_value_col:
            cell.alignment = HEADER_CENTER


def prepare_header_row(ws, row):
    """Clear label cells (B, C) and apply header styling for standard sheets."""
    clear_header_cells(ws, row)
    style_header_row(ws, row, first_value_col=STANDARD_FIRST_VAL_COL, last_col=ws.max_column)


def style_title(ws, row, col=2):
    """Apply title style: 12pt bold Calibri, no fill, no border."""
    cell = ws.cell(row=row, column=col)
    cell.font = TITLE_FONT
    cell.fill = NO_FILL
    cell.border = NO_BORDER


# ── Bold helpers ─────────────────────────────────────────────────────────────

def bold_total_column(ws, header_text=TOTAL_COL):
    """Bold all cells in the column whose header matches header_text."""
    total_col = None
    for col in range(COL_B, ws.max_column + 1):
        if ws.cell(row=HEADER_ROW, column=col).value == header_text:
            total_col = col
            break
    if total_col is None:
        return
    for row in range(HEADER_ROW, ws.max_row + 1):
        ws.cell(row=row, column=total_col).font = BOLD


def bold_row(ws, row):
    """Bold every cell in the given row (columns B through max)."""
    for col in range(COL_B, ws.max_column + 1):
        ws.cell(row=row, column=col).font = BOLD


def bold_total_rows(ws):
    """Bold every row whose column B or C contains 'TOTAL'."""
    for row in range(1, ws.max_row + 1):
        for check_col in (COL_B, COL_C):
            if ws.cell(row=row, column=check_col).value == TOTAL_COL:
                bold_row(ws, row)
                break


# ── Global formatting passes ────────────────────────────────────────────────

def apply_number_format(writer):
    """Apply #,##0;(#,##0);- to all numeric cells (no decimals, parentheses for negatives)."""
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        for row in ws.iter_rows(min_row=EXCEL_DATA_ROW, min_col=3):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = NUM_FMT


def highlight_excluded_rows(writer, excluded, sheet_name):
    ws = writer.sheets[sheet_name]
    for row in range(EXCEL_DATA_ROW, ws.max_row + 1):
        cuenta = ws.cell(row=row, column=COL_B).value
        if cuenta in excluded:
            for col in range(COL_B, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = HIGHLIGHT_FILL


def highlight_undefined_bs_rows(ws, undefined_cuentas):
    """Highlight rows whose CUENTA_CONTABLE is in the undefined set (POR DEFINIR) with red."""
    if not undefined_cuentas:
        return
    for row in range(EXCEL_DATA_ROW, ws.max_row + 1):
        cuenta = ws.cell(row=row, column=COL_B).value
        if cuenta in undefined_cuentas:
            for col in range(COL_B, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = UNDEFINED_FILL
                cell.font = UNDEFINED_FONT


def format_subtotal_rows(writer, sheet_name, label_set):
    """Style subtotal rows matching *label_set* with purple fill + bold + dotted border."""
    ws = writer.sheets[sheet_name]
    for row in range(EXCEL_DATA_ROW, ws.max_row + 1):
        label = ws.cell(row=row, column=COL_B).value
        if label in label_set:
            for col in range(COL_B, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = SUBTOTAL_FILL
                cell.font = BOLD
                cell.border = DOTTED_BORDER


# ── Per-sheet styling ────────────────────────────────────────────────────────

def style_sheet(ws, layout_type="standard"):
    """Apply standard styling to a sheet based on layout type.

    layout_type:
      "standard"       — standard col widths, clear label headers, bold TOTAL col+rows
      "standard_raw"   — standard col widths, raw header styling (no label clear), bold TOTAL col
      "wide"           — wide col widths, raw header styling, bold TOTAL col
    """
    if layout_type == "wide":
        _set_wide_col_widths(ws)
        style_header_row(ws, HEADER_ROW, first_value_col=WIDE_FIRST_VAL_COL, last_col=ws.max_column)
    else:
        set_standard_col_widths(ws)
        if layout_type == "standard":
            prepare_header_row(ws, HEADER_ROW)
        else:
            style_header_row(ws, HEADER_ROW, first_value_col=STANDARD_FIRST_VAL_COL, last_col=ws.max_column)

    bold_total_column(ws)

    if layout_type == "standard":
        bold_total_rows(ws)


def style_validator_sheets(writer):
    for sheet_name in ("VALIDADOR1", "VALIDADOR2"):
        style_sheet(writer.sheets[sheet_name], "standard_raw")
    style_sheet(writer.sheets["VALIDADOR3"], "wide")


def style_detail_sheets(writer):
    """Style COSTO_DETAIL, GASTO_VENTA_DETAIL, GASTO_ADMIN_DETAIL, OTROS_EGRESOS_DETAIL."""
    for sheet_name in ("COSTO_DETAIL", "GASTO_VENTA_DETAIL", "GASTO_ADMIN_DETAIL", "OTROS_EGRESOS_DETAIL"):
        style_sheet(writer.sheets[sheet_name], "standard")


def style_detalle_cc_x_cc(writer):
    """Style DETALLE_CC_x_CC: wide layout, section titles, TOTAL column bold per table."""
    ws = writer.sheets["DETALLE_CC_x_CC"]
    _set_wide_col_widths(ws)
    month_abbrevs = MONTH_NAMES_SET

    for row in range(1, ws.max_row + 1):
        # Style section title rows (e.g. "xx. DETALLE DE COSTOS")
        val_b = ws.cell(row=row, column=COL_B).value
        if isinstance(val_b, str) and val_b.startswith("xx."):
            style_title(ws, row)

        # Style header rows (rows with month abbreviations like JAN, FEB, etc.)
        val_f = ws.cell(row=row, column=WIDE_FIRST_VAL_COL).value
        if val_f in month_abbrevs:
            style_header_row(ws, row, first_value_col=WIDE_FIRST_VAL_COL, last_col=ws.max_column)

        # Bold TOTAL column (R) for data
        cell_r = ws.cell(row=row, column=WIDE_TOTAL_COL)
        if cell_r.value is not None and row > HEADER_ROW:
            cell_r.font = BOLD


def style_sales_details(writer, sheet_name=None):
    """Style the Ingresos + Proyectos sheet: widths, headers, bold TOTAL col + rows."""
    if sheet_name is None:
        # Legacy fallback
        sheet_name = "SALES DETAILS"
    ws = writer.sheets[sheet_name]
    set_standard_col_widths(ws)

    prepare_header_row(ws, SALES_HEADER_ROW)

    # Find and style ALL sub-section title + header rows
    for row in range(SALES_HEADER_ROW + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=COL_B).value
        if isinstance(val, str) and (val.startswith("Nota") or val.startswith("Proyectos") or val.startswith("Ingresos")):
            style_title(ws, row)
            prepare_header_row(ws, row + 1)

    bold_total_column(ws)
    bold_total_rows(ws)


def style_pl_sheet(writer):
    ws = writer.sheets["PL"]
    for col_idx in range(COL_C, PL_TOTAL_COL + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = VAL_COL_WIDTH

    # Remove PARTIDA_PL header, style value headers
    _clear_header_cell(ws, HEADER_ROW, COL_B)
    # PL value columns start at C
    style_header_row(ws, HEADER_ROW, first_value_col=COL_C, last_col=ws.max_column)

    # Bold TOTAL column (O)
    for row in range(HEADER_ROW, ws.max_row + 1):
        ws.cell(row=row, column=PL_TOTAL_COL).font = BOLD


# ── BS styling ───────────────────────────────────────────────────────────

def format_bs_partida_rows(writer, sheet_name="BS"):
    """Bold the partida header rows (the classification labels with their subtotals)."""
    ws = writer.sheets[sheet_name]
    for row in range(EXCEL_DATA_ROW, ws.max_row + 1):
        label = ws.cell(row=row, column=COL_B).value
        if label in BS_PARTIDA_LABELS:
            for col in range(COL_B, ws.max_column + 1):
                ws.cell(row=row, column=col).font = BOLD


def style_bs_sheet(writer, sheet_name="BS"):
    ws = writer.sheets[sheet_name]
    for col_idx in range(COL_C, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = BS_VAL_COL_WIDTH
    _clear_header_cell(ws, HEADER_ROW, COL_B)
    style_header_row(ws, HEADER_ROW, first_value_col=COL_C, last_col=ws.max_column)


def style_bs_detail_sheets(writer):
    """Style all single-table BS detail sheets with standard widths, headers, bold TOTAL."""
    for sheet_name in ("DETALLE_EFECTIVO", "DETALLE_OTROS_ACTIVOS", "DETALLE_TRIBUTOS", "DETALLE_OBLIGACIONES"):
        if sheet_name not in writer.sheets:
            continue
        style_sheet(writer.sheets[sheet_name], "standard")


# ── Init + multi-table styling ──────────────────────────────────────────────

def style_init(ws, col_b_width=COL_B_WIDTH_DEFAULT):
    """Set column A/B widths and style the title row (row 2)."""
    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.column_dimensions["B"].width = col_b_width
    style_title(ws, TITLE_ROW)


def style_two_table_sheet(ws):
    """Style a multi-table sheet: section titles, headers, bold TOTAL rows/col, standard widths.

    Consolidates bold_total_column, bold_total_rows, and header detection
    into a single pass over the sheet rows for performance.
    """
    set_standard_col_widths(ws)

    # Locate the TOTAL column from the first header row (header-only scan)
    total_col = None
    for col in range(COL_B, ws.max_column + 1):
        if ws.cell(row=HEADER_ROW, column=col).value == TOTAL_COL:
            total_col = col
            break

    month_abbrevs = MONTH_NAMES_SET

    # Single pass: bold TOTAL col cells, bold TOTAL rows, detect headers
    for row in range(1, ws.max_row + 1):
        # Bold the TOTAL column cell
        if total_col is not None and row >= HEADER_ROW:
            ws.cell(row=row, column=total_col).font = BOLD

        # Bold entire row if column B or C is "TOTAL"
        for check_col in (COL_B, COL_C):
            if ws.cell(row=row, column=check_col).value == TOTAL_COL:
                bold_row(ws, row)
                break

        # Detect month-header rows
        val = ws.cell(row=row, column=STANDARD_FIRST_VAL_COL).value
        if val in month_abbrevs:
            prepare_header_row(ws, row)
            title_row = row - 2
            if title_row >= 1 and ws.cell(row=title_row, column=COL_B).value:
                style_title(ws, title_row)
        # NIT pivot header: clear NIT/RAZON_SOCIAL labels, style remaining cols
        elif (ws.cell(row=row, column=COL_B).value == "NIT"
              and ws.cell(row=row, column=COL_C).value == "RAZON_SOCIAL"):
            clear_header_cells(ws, row)
            style_header_row(ws, row, first_value_col=STANDARD_FIRST_VAL_COL,
                             last_col=ws.max_column)
            title_row = row - 2
            if title_row >= 1 and ws.cell(row=title_row, column=COL_B).value:
                style_title(ws, title_row)
