import os
import openpyxl
from openpyxl.styles import PatternFill
import numpy_financial as npf
import numpy as np
import pandas as pd
import pyodbc
import win32com.client as win32
from datetime import datetime
import shutil
import re
import config


# -----------------------------------------------------------------------------
# STEP 2: FUNCTIONS
# -----------------------------------------------------------------------------

def create_results_dataframe(results_list):
    """
    Converts a list of dictionaries into a pandas DataFrame.
    """
    if not results_list:
        return pd.DataFrame()
    return pd.DataFrame(results_list)


def get_validation_list(db_path, table_name, column_name):
    """Fetches a list of unique values from an Access database table."""
    conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + db_path + r';'
    cnxn = pyodbc.connect(conn_str)
    cursor = cnxn.cursor()
    sql = f"SELECT [{column_name}] FROM [{table_name}]"
    rows = cursor.execute(sql).fetchall()
    valid_values = {row[0] for row in rows}
    cursor.close()
    cnxn.close()
    return valid_values


def validate_excel_data(file_path, valid_servicios, valid_cotizaciones, red_fill):
    """
    Validates data in the single PLANTILLA sheet against the FINAL rules.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    plantilla_sheet = wb[config.PLANTILLA_SHEET_NAME]
    file_is_clean = True

    start_row = config.DETAIL_START_ROW
    end_row = plantilla_sheet.max_row + 1

    VALID_CATEGORIES = {"FACTIBILIDAD", "EQUIPAMIENTO", "INVERSIÓN", "INVERSION"}
    TRIGGER_CATEGORIES = {"EQUIPAMIENTO"}

    TIPO_SERVICIO_COL = config.FIXED_TIPO_SERVICIO_COL
    TICKET_COL = config.FIXED_TICKET_COL
    CATEGORY_COL = config.FIXED_CATEGORY_COL
    RECURRING_COL = config.RECURRING_SERVICE_COL

    for row in range(start_row, end_row):
        cell = plantilla_sheet[f'{RECURRING_COL}{row}']
        if cell.value and cell.value not in valid_servicios:
            cell.fill = red_fill
            file_is_clean = False

    for row in range(start_row, end_row):
        c_cell = plantilla_sheet[f'{CATEGORY_COL}{row}']
        d_cell = plantilla_sheet[f'{TIPO_SERVICIO_COL}{row}']
        e_cell = plantilla_sheet[f'{TICKET_COL}{row}']
        c_value = str(c_cell.value).strip().upper() if c_cell.value else None

        if c_cell.value is None and d_cell.value is None and e_cell.value is None:
            break

        if c_value and c_value not in VALID_CATEGORIES:
            c_cell.fill = red_fill
            file_is_clean = False

        if d_cell.value and d_cell.value not in valid_cotizaciones:
            d_cell.fill = red_fill
            file_is_clean = False

        if c_value in TRIGGER_CATEGORIES:
            e_value_is_empty = (e_cell.value is None or str(e_cell.value).strip() == "")
            if e_value_is_empty:
                e_cell.fill = red_fill
                file_is_clean = False
    return file_is_clean, wb


def calculate_financial_metrics(**kwargs):
    """
    Calculates key financial metrics (VAN, TIR, Payback) based on cash flows.
    """
    costoCapital_annual = kwargs.get('costoCapital_annual', 0)
    plazoContrato = int(kwargs.get('plazoContrato', 0))
    MRC = kwargs.get('MRC', 0)
    NRC = kwargs.get('NRC', 0)
    comisionesPersonal = kwargs.get('comisionesPersonal', 0)
    costoInstalacion = kwargs.get('costoInstalacion', 0)
    tipoCambio = kwargs.get('tipoCambio', 1)
    total_monthly_expense = kwargs.get('total_monthly_expense', 0)

    # --- Totals Calculation ---
    total_revenues_calculated = NRC + (MRC * plazoContrato)

    upfront_costs = (costoInstalacion * tipoCambio) + comisionesPersonal
    total_monthly_expense_converted = total_monthly_expense * tipoCambio
    total_expenses_calculated = upfront_costs + (total_monthly_expense_converted * plazoContrato)

    gross_margin = total_revenues_calculated - total_expenses_calculated

    # --- Cash Flow Calculation for NPV, IRR, Payback ---
    cash_flows = []
    payback_period = None
    cumulative_cash_flow = 0

    period_0_cash_flow = -upfront_costs
    cash_flows.append(period_0_cash_flow)
    cumulative_cash_flow += period_0_cash_flow

    for period in range(1, plazoContrato + 1):
        monthly_cash_flow = MRC - total_monthly_expense_converted
        if period == 1:
            monthly_cash_flow += NRC

        cash_flows.append(monthly_cash_flow)
        cumulative_cash_flow += monthly_cash_flow

        if payback_period is None and cumulative_cash_flow >= 0:
            payback_period = period

    npv = npf.npv(costoCapital_annual / 12, cash_flows)

    try:
        irr = npf.irr(cash_flows)
        final_TIR = f"{irr:.6f}" if np.isfinite(irr) else 'N/A'
    except Exception:
        final_TIR = 'N/A'

    return {
        'totalRevenue': total_revenues_calculated,
        'totalExpense': total_expenses_calculated,
        'paybackPeriod': payback_period,
        'grossMargin': gross_margin,
        'TIR': final_TIR,
        'VAN': npv
    }


def append_to_access_database_minimized(header_df, fixed_costs_df, recurring_df, db_path):
    """
    Appends three pandas DataFrames to their respective tables in the Access database.
    """
    conn_str = (r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + db_path + r';')
    cnxn = pyodbc.connect(conn_str)
    cursor = cnxn.cursor()

    tables_to_insert = [
        (config.HEADER_TABLE_NAME, header_df),
        (config.FIXED_COST_TABLE_NAME, fixed_costs_df),
        (config.RECURRING_TABLE_NAME, recurring_df),
    ]

    for table_name, df in tables_to_insert:
        if df.empty:
            continue

        table_columns = [row.column_name for row in cursor.columns(table=table_name)]
        df_filtered = df[[col for col in df.columns if col in table_columns]]

        columns = ', '.join(f"[{col}]" for col in df_filtered.columns)
        placeholders = ', '.join('?' * len(df_filtered.columns))
        sql = f"INSERT INTO [{table_name}] ({columns}) VALUES ({placeholders})"

        for _, row in df_filtered.iterrows():
            try:
                values = [None if pd.isna(val) else val for val in row]
                cursor.execute(sql, values)
            except pyodbc.Error as e:
                print(f"Error inserting row into [{table_name}]: {e}")
                print(f"Problematic row data: {row.to_dict()}")

    cnxn.commit()
    cursor.close()
    cnxn.close()


def move_file_to_historico(file_path, base_dest_path, unidad_negocio_value, client_name, submission_timestamp):
    """
    Moves a processed file and returns the new destination path.
    """
    safe_folder_name = re.sub(r'[^\w\s-]', '_', str(unidad_negocio_value)).strip().upper()
    clean_client_name = re.sub(r'[\s]+', '_', str(client_name)).strip('_')
    clean_client_name = re.sub(r'[^\w_]', '', clean_client_name)

    try:
        if not isinstance(submission_timestamp, str):
            submission_timestamp = submission_timestamp.strftime('%Y-%m-%d %H:%M:%S')
        dt_object = datetime.strptime(submission_timestamp, "%Y-%m-%d %H:%M:%S")
        date_suffix = dt_object.strftime("%Y_%m_%d")
    except (ValueError, TypeError):
        date_suffix = datetime.now().strftime("%Y_%m_%d")

    base_file_name = f"{clean_client_name}_{date_suffix}"
    extension = ".xlsx"
    new_file_name = f"{base_file_name}{extension}"
    dest_folder = os.path.join(base_dest_path, safe_folder_name)

    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)

    counter = 1
    final_dest_path = os.path.join(dest_folder, new_file_name)

    while os.path.exists(final_dest_path):
        new_file_name = f"{base_file_name}_{counter}{extension}"
        final_dest_path = os.path.join(dest_folder, new_file_name)
        counter += 1

    try:
        shutil.move(file_path, final_dest_path)
        print(f"File successfully archived as: {new_file_name} in {dest_folder}")
        return final_dest_path
    except Exception as e:
        print(f"Error moving file '{os.path.basename(file_path)}': {e}")
        return None


def get_salesman_email(db_path, salesman_name):
    """
    Fetches the email address for a given salesman from the Access database.
    """
    DEFAULT_EMAIL = "sternero@fiberlux.pe"
    conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + db_path + r';'
    cnxn = None
    try:
        cnxn = pyodbc.connect(conn_str)
        cursor = cnxn.cursor()
        sql = f"SELECT [{config.SALESMAN_EMAIL_COLUMN}] FROM [{config.SALESMAN_TABLE_NAME}] WHERE [{config.SALESMAN_NAME_COLUMN}] = ?"
        cursor.execute(sql, (salesman_name,))
        row = cursor.fetchone()
        return str(row[0]).strip() if row else DEFAULT_EMAIL
    except pyodbc.Error as e:
        print(f"Database error looking up email: {e}")
        return DEFAULT_EMAIL
    finally:
        if cnxn:
            cnxn.close()


def send_outlook_email(recipient, subject, body, attachment_path=None):
    """
    Sends an email using the local Outlook application via win32com.
    """
    try:
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = recipient
        mail.Subject = subject
        mail.HTMLBody = f"<html><body>{body}</body></html>"
        if attachment_path and os.path.exists(attachment_path):
            mail.Attachments.Add(attachment_path)
        mail.Send()
        print(f"Email successfully sent to {recipient}.")
    except Exception as e:
        print(f"Failed to send email to {recipient}. Ensure Outlook is open. Error: {e}")


def calculate_state_commission_automated(first_metrics, variables_to_pass, is_pago_unico):
    """
    Handles the conditional commission calculation for 'ESTADO' non-interactively.
    """
    total_revenues = first_metrics.get('totalRevenue', 0)
    gross_margin = first_metrics.get('grossMargin', 0)

    if total_revenues == 0:
        return {'comisionesPersonal': 0, 'commission_rate': 0, **first_metrics}

    rentabilidad_sin_comision = gross_margin / total_revenues
    final_commission_amount = 0.0
    commission_rate = 0.0

    if is_pago_unico:
        limit_pen = 0.0
        if 0.30 <= rentabilidad_sin_comision <= 0.35:
            commission_rate, limit_pen = 0.01, 11000
        elif 0.35 < rentabilidad_sin_comision <= 0.39:
            commission_rate, limit_pen = 0.02, 12000
        elif 0.39 < rentabilidad_sin_comision <= 0.49:
            commission_rate, limit_pen = 0.03, 13000
        elif 0.49 < rentabilidad_sin_comision <= 0.59:
            commission_rate, limit_pen = 0.04, 14000
        elif rentabilidad_sin_comision > 0.59:
            commission_rate, limit_pen = 0.05, 15000

        if commission_rate > 0:
            calculated_commission = total_revenues * commission_rate
            final_commission_amount = min(calculated_commission, limit_pen)
    else:
        plazo = first_metrics.get('plazoContrato', 0)
        payback = first_metrics.get('paybackPeriod')
        mrc = first_metrics.get('MRC', 0)
        payback_ok = (payback is not None)
        limit_mrc_multiplier = 0.0

        if plazo == 12:
            if 0.30 <= rentabilidad_sin_comision <= 0.35 and payback_ok and payback <= 7:
                commission_rate, limit_mrc_multiplier = 0.025, 0.8
            elif 0.35 < rentabilidad_sin_comision <= 0.39 and payback_ok and payback <= 7:
                commission_rate, limit_mrc_multiplier = 0.03, 0.9
            elif rentabilidad_sin_comision > 0.39 and payback_ok and payback <= 6:
                commission_rate, limit_mrc_multiplier = 0.035, 1.0
        elif plazo == 24:
            if 0.30 <= rentabilidad_sin_comision <= 0.35 and payback_ok and payback <= 11:
                commission_rate, limit_mrc_multiplier = 0.025, 0.8
            elif 0.35 < rentabilidad_sin_comision <= 0.39 and payback_ok and payback <= 11:
                commission_rate, limit_mrc_multiplier = 0.03, 0.9
            elif rentabilidad_sin_comision > 0.39 and payback_ok and payback <= 10:
                commission_rate, limit_mrc_multiplier = 0.035, 1.0
        elif plazo == 36:
            if 0.30 <= rentabilidad_sin_comision <= 0.35 and payback_ok and payback <= 19:
                commission_rate, limit_mrc_multiplier = 0.025, 0.8
            elif 0.35 < rentabilidad_sin_comision <= 0.39 and payback_ok and payback <= 19:
                commission_rate, limit_mrc_multiplier = 0.03, 0.9
            elif rentabilidad_sin_comision > 0.39 and payback_ok and payback <= 18:
                commission_rate, limit_mrc_multiplier = 0.035, 1.0
        elif plazo == 48:
            if 0.30 <= rentabilidad_sin_comision <= 0.35 and payback_ok and payback <= 26:
                commission_rate, limit_mrc_multiplier = 0.02, 0.8
            elif 0.35 < rentabilidad_sin_comision <= 0.39 and payback_ok and payback <= 26:
                commission_rate, limit_mrc_multiplier = 0.025, 0.9
            elif rentabilidad_sin_comision > 0.39 and payback_ok and payback <= 25:
                commission_rate, limit_mrc_multiplier = 0.03, 1.0

        if commission_rate > 0.0:
            calculated_commission = total_revenues * commission_rate
            limit_mrc_amount = mrc * limit_mrc_multiplier
            final_commission_amount = min(calculated_commission, limit_mrc_amount)

    variables_to_pass['comisionesPersonal'] = final_commission_amount
    final_metrics = calculate_financial_metrics(**variables_to_pass)

    final_metrics['comisionesPersonal'] = final_commission_amount
    final_metrics['commission_rate'] = commission_rate
    return final_metrics


def get_gigalan_inputs():
    """
    Prompts the user for GIGALAN-specific information.
    """
    while True:
        print("\n--- Categorización GIGALAN ---")
        print("Regiones válidas: 'LIMA', 'PROVINCIAS CON CACHING', 'PROVINCIAS CON INTERNEXA', 'PROVINCIAS CON TDP'")
        region_input = input("Ingrese la región exacta del proyecto: ").strip().upper()
        valid_regions = {'LIMA', 'PROVINCIAS CON CACHING', 'PROVINCIAS CON INTERNEXA', 'PROVINCIAS CON TDP'}
        if region_input in valid_regions:
            region = region_input
            break
        else:
            print("Entrada inválida.")

    while True:
        project_type_input = input("¿Es una venta nueva o un upgrade? (N/U): ").strip().upper()
        if project_type_input == 'N':
            project_type = 'VENTA NUEVA'
            old_mrc = 0.0
            break
        elif project_type_input == 'U':
            project_type = 'UPGRADE'
            while True:
                try:
                    old_mrc = float(input("Ingrese el valor del MRC antiguo: "))
                    break
                except ValueError:
                    print("Entrada inválida. Ingrese un número.")
            break
        else:
            print("Entrada inválida. Ingrese 'N' o 'U'.")
    return region, project_type, old_mrc


# ## KEY FIX HERE: Restored the full, original commission logic for GIGALAN ##
def calculate_gigalan_commission(base_metrics, region, project_type, old_mrc):
    """
    Calculates the GIGALAN commission based on the defined rules.
    """
    payback = base_metrics.get('paybackPeriod')
    # Use a safe default of 1 for totalRevenue to avoid division by zero
    rentabilidad = base_metrics.get('grossMargin', 0) / base_metrics.get('totalRevenue', 1)

    if payback is None or payback > 2:
        print("\nADVERTENCIA: Payback > 2 meses. No aplica comisión.")
        return 0.0, 0.0

    commission_rate = 0.0
    plazo = base_metrics.get('plazoContrato', 0)
    mrc = base_metrics.get('MRC', 0)

    # --- FULL GIGALAN COMMISSION LOGIC ---
    if region == 'LIMA':
        if project_type == 'VENTA NUEVA':
            if 0.40 <= rentabilidad < 0.50:
                commission_rate = 0.009
            elif 0.50 <= rentabilidad < 0.60:
                commission_rate = 0.014
            elif 0.60 <= rentabilidad < 0.70:
                commission_rate = 0.019
            elif rentabilidad >= 0.70:
                commission_rate = 0.024
        elif project_type == 'UPGRADE':
            if 0.40 <= rentabilidad < 0.50:
                commission_rate = 0.01
            elif 0.50 <= rentabilidad < 0.60:
                commission_rate = 0.015
            elif 0.60 <= rentabilidad < 0.70:
                commission_rate = 0.02
            elif rentabilidad >= 0.70:
                commission_rate = 0.025

    elif region == 'PROVINCIAS CON CACHING':
        if 0.40 <= rentabilidad < 0.45:
            commission_rate = 0.03
        elif rentabilidad >= 0.45:
            commission_rate = 0.035

    elif region == 'PROVINCIAS CON INTERNEXA':
        if 0.17 <= rentabilidad < 0.20:
            commission_rate = 0.02
        elif rentabilidad >= 0.20:
            commission_rate = 0.03

    elif region == 'PROVINCIAS CON TDP':
        if 0.17 <= rentabilidad < 0.20:
            commission_rate = 0.02
        elif rentabilidad >= 0.20:
            commission_rate = 0.03

    # --- FINAL CALCULATION ---
    if project_type == 'VENTA NUEVA':
        calculated_commission = commission_rate * mrc * plazo
    elif project_type == 'UPGRADE':
        calculated_commission = commission_rate * plazo * (mrc - old_mrc)
    else:
        calculated_commission = 0.0

    return calculated_commission, commission_rate


def extract_and_calculate_fixed_costs(plantilla_sheet, transaction_id):
    fixed_details_list = []
    total_costoInstalacion = 0.0
    start_row = config.DETAIL_START_ROW
    end_row = plantilla_sheet.max_row + 1

    for row in range(start_row, end_row):
        categoria = plantilla_sheet[f'{config.FIXED_CATEGORY_COL}{row}'].value
        total = get_numeric_value(plantilla_sheet[f'{config.FIXED_COST_TOTAL_COL}{row}'].value)

        if categoria is None and total == 0.0:
            break

        total_costoInstalacion += total
        fixed_details_list.append({
            'transactionID': transaction_id,
            'categoria': categoria,
            'tipo_servicio': plantilla_sheet[f'{config.FIXED_TIPO_SERVICIO_COL}{row}'].value,
            'ticket': plantilla_sheet[f'{config.FIXED_TICKET_COL}{row}'].value,
            'ubicacion': plantilla_sheet[f'{config.FIXED_UBICACION_COL}{row}'].value,
            'cantidad': get_numeric_value(plantilla_sheet[f'{config.FIXED_Q_COL}{row}'].value),
            'costoUnitario': get_numeric_value(plantilla_sheet[f'{config.FIXED_COST_UNITARIO_COL}{row}'].value),
            'total': total
        })
    return fixed_details_list, total_costoInstalacion


def extract_and_calculate_recurring_services(plantilla_sheet, transaction_id):
    recurring_details_list = []
    total_monthly_expense = 0.0
    start_row = config.DETAIL_START_ROW
    end_row = plantilla_sheet.max_row + 1

    for row in range(start_row, end_row):
        tipo_servicio = plantilla_sheet[f'{config.RECURRING_SERVICE_COL}{row}'].value
        cantidad_Q = get_numeric_value(plantilla_sheet[f'{config.RECURRING_Q_COL}{row}'].value)
        precio_unitario_P = get_numeric_value(plantilla_sheet[f'{config.RECURRING_REVENUE_UNIT_COL}{row}'].value)

        if tipo_servicio is None and cantidad_Q == 0.0 and precio_unitario_P == 0.0:
            break

        costo_unitario_1 = get_numeric_value(plantilla_sheet[f'{config.RECURRING_COST_1_COL}{row}'].value)
        costo_unitario_2 = get_numeric_value(plantilla_sheet[f'{config.RECURRING_COST_2_COL}{row}'].value)
        ingreso_calculado = cantidad_Q * precio_unitario_P
        egreso_calculado = (costo_unitario_1 + costo_unitario_2) * cantidad_Q
        total_monthly_expense += egreso_calculado

        recurring_details_list.append({
            'transactionID': transaction_id,
            'tipo_servicio': tipo_servicio,
            'nota': plantilla_sheet[f'{config.RECURRING_NOTA_COL}{row}'].value,
            'ubicacion': plantilla_sheet[f'{config.RECURRING_UBICACION_COL}{row}'].value,
            'Q': cantidad_Q,
            'P': precio_unitario_P,
            'ingreso': ingreso_calculado,
            'CU1': costo_unitario_1,
            'CU2': costo_unitario_2,
            'proveedor': plantilla_sheet[f'{config.RECURRING_PROVEEDOR_COL}{row}'].value,
            'egreso': egreso_calculado
        })
    return recurring_details_list, total_monthly_expense


def get_numeric_value(cell_value):
    """
    Safely converts a cell value to a float, treating None/empty as 0.0.
    """
    if cell_value is None or cell_value == "":
        return 0.0
    try:
        return float(cell_value)
    except (ValueError, TypeError):
        return 0.0