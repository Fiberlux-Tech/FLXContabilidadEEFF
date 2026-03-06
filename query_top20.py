"""
Top-20 RAZON_SOCIAL by revenue (70.xxxx) and cost (60.xxxx)
Period: January–October 2025, one pair of tables per company.
Output: query_top20_70_60.xlsx
"""

import logging
from datetime import date

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

from data.db import connect
from data.queries import SQL_SCHEMA, SQL_VIEW
from config.company import VALID_COMPANIES

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

YEAR = 2025
START = date(2025, 1, 1)
END = date(2025, 11, 1)  # up to end of October

MONTH_COLS = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
              "JUL", "AGO", "SEP", "OCT"]
MONTH_MAP = {i: name for i, name in enumerate(MONTH_COLS, start=1)}


def fetch_prefix(conn, company: str, prefix: str) -> pd.DataFrame:
    """Fetch rows for a single account prefix (e.g. '70' or '60')."""
    query = (
        "SELECT CIA, CUENTA_CONTABLE, RAZON_SOCIAL, FECHA, "
        "DEBITO_LOCAL, CREDITO_LOCAL "
        f"FROM {SQL_SCHEMA}.{SQL_VIEW} "
        "WHERE CIA = ? AND FECHA >= ? AND FECHA < ? "
        "AND CUENTA_CONTABLE LIKE ? "
        "AND FUENTE NOT LIKE 'CIERRE%'"
    )
    params = [company, START, END, f"{prefix}%"]
    return pd.read_sql(query, conn, params=params)


def build_top20(df: pd.DataFrame, sign: int) -> pd.DataFrame:
    """
    Group by RAZON_SOCIAL, pivot by month, keep top 20 by absolute total.
    sign: +1 for revenue (CREDITO-DEBITO), -1 to flip (DEBITO-CREDITO for costs).
    """
    if df.empty:
        return pd.DataFrame(columns=["RAZON_SOCIAL"] + MONTH_COLS + ["TOTAL"])

    df = df.copy()
    df["FECHA"] = pd.to_datetime(df["FECHA"])
    df["MES"] = df["FECHA"].dt.month
    df["SALDO"] = sign * (df["CREDITO_LOCAL"] - df["DEBITO_LOCAL"])

    grouped = (
        df.groupby(["RAZON_SOCIAL", "MES"], as_index=False)["SALDO"]
        .sum()
    )

    pivot = grouped.pivot_table(
        index="RAZON_SOCIAL", columns="MES", values="SALDO",
        aggfunc="sum", fill_value=0,
    )

    # Ensure all months 1-10 present
    for m in range(1, 11):
        if m not in pivot.columns:
            pivot[m] = 0.0
    pivot = pivot[[m for m in range(1, 11)]]
    pivot.columns = MONTH_COLS

    pivot["TOTAL"] = pivot[MONTH_COLS].sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False).head(20)
    pivot = pivot.reset_index()
    return pivot


# ── Excel styling helpers ────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
BODY_FONT = Font(name="Calibri", size=10)
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)


def write_table(ws, start_row: int, title: str, df: pd.DataFrame) -> int:
    """Write a titled table to the worksheet, return the next free row."""
    # Title
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    start_row += 1

    if df.empty:
        ws.cell(row=start_row, column=1, value="Sin datos").font = BODY_FONT
        return start_row + 2

    cols = list(df.columns)

    # Header row
    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    for r_offset, (_, row) in enumerate(df.iterrows(), start=1):
        r = start_row + r_offset
        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c_idx, value=row[col_name])
            cell.border = THIN_BORDER
            if col_name == "RAZON_SOCIAL":
                cell.font = BODY_FONT
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.font = BODY_FONT
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            # Highlight TOTAL column
            if col_name == "TOTAL":
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL

    # Column widths
    ws.column_dimensions["A"].width = 42
    for c_idx in range(2, len(cols) + 1):
        ws.column_dimensions[chr(64 + c_idx)].width = 14

    return start_row + len(df) + 2  # +2 for spacing


def main():
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    with connect() as conn:
        for company in sorted(VALID_COMPANIES):
            logger.info("Processing %s ...", company)

            df_70 = fetch_prefix(conn, company, "70")
            df_60 = fetch_prefix(conn, company, "60")

            top20_70 = build_top20(df_70, sign=+1)   # revenue: positive
            top20_60 = build_top20(df_60, sign=-1)    # cost: show as positive

            ws = wb.create_sheet(title=company[:31])
            row = 1
            row = write_table(ws, row, f"{company} — Top 20 Ingresos (70.xxxx)", top20_70)
            row += 1
            write_table(ws, row, f"{company} — Top 20 Costos (60.xxxx)", top20_60)

    out = "query_top20_70_60.xlsx"
    wb.save(out)
    logger.info("Saved → %s", out)


if __name__ == "__main__":
    main()
